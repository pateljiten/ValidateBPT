from openpyxl import Workbook
from openpyxl import load_workbook
import subprocess

logFile = open("ValidateBPT_data.log","w");
bl_wb = load_workbook(filename='CNT_Billing_BPT_GE_INV_V06_CNT_20160629.xlsx',read_only=True);
bl_sheets = ('Charge Codes','Bill Cycles','Messages Entities','Quality Assurance');
cm_wb = load_workbook(filename='CNT_Billing_BPT_GE_CM_V04_AMD_20160629.xlsx',read_only=True);
cm_sheets = ('CUSTOMER_TYPE','CUSTOMER_SUB_TYPES','CSM_PYM_METHOD_VALID','CM1_CUST_TP_PYM_DUE_DAY','GN1_CURRENCY','CSM_CYCLE','CM1_CYCLE_PARTITION_REL','CM1_GENERIC_CODES');
ar_wb = load_workbook(filename='CNT_Billing_BPT_GE_AR_V09_CNT_20160629.xlsx',read_only=True);
ar_sheets = ('Payment','Backout','Funds Transfer','Refunds','Credit Reason','Direct Debit','Add RT','AR1_CHARGE_CODE','AR1_GENERIC_CODES');
cl_wb = load_workbook(filename='CNT_Billing_BPT_GE_CL_V02_CNT_20160629.xlsx',read_only=True);
cl_sheets = ('CL1_CATEGORY','CL1_ATTR_MAPPING','CL1_COLLECTOR_GROUP','CL1_AGENCY','CL1_ACTIVITY','CL1_ACT_RSN','CL1_FOLLOW_UP','CL1_FOLLOW_UP','CL1_AR_FINANCE_ACTIVITY' \
,'UTL1_BD_HOLIDAYS','CL1_AR_DEBT_TYPE','CL1_CUST_COMUNICAT','CL1_GENERIC_CODES','UTL1_BD_CALENDARS','CL1_PUNISHMENT_LEVEL','CL1_PROPERTIES'\
,'CL1_PA_POLICY','CL1_BPM_PROCESS','CCS1_REQ_TYPE','CCS1_REQ_TYPE_RESOURCE','CCS1_TEMPLATE','CCS1_COMMUNICATION_TYPE','CCS1_RESOURCE','CCS1_APPLICATION' \
,'CCS1_ENTITY_TYPE','CCS1_ACTIVITY','CCS1_ACTIVITY_RESULT');
bptDatList = []
bpt_config = {}

def log(msg):
	global logFile
	logFile.write("----------------------------------------------\n");
	logFile.write(msg);
	logFile.write("\n----------------------------------------------\n");

def init():
	global bpt_config;
	log("Loading Configuration .....");
	#SHEET_NAME|RANGE|START_ROW|START_COL|END_ROW|END_COL
	with open("BPTConfig.txt","r") as f:
		for line_terminated in f:
			line = line_terminated.rstrip('\n')			
			tokens = line.split("|")
			if(tokens[0] in bpt_config):
				bpt_config[tokens[0]].update({tokens[1]:(int(tokens[2]),int(tokens[3]),int(tokens[4]),int(tokens[5]))})		
			else:
				bpt_config[tokens[0]] = {tokens[1]:(int(tokens[2]),int(tokens[3]),int(tokens[4]),int(tokens[5]))}					
	log("Loading Configuration Complete .....");
	#print (bpt_config);
	
def extract(wb,sheetList):	
	log("Extracting BPT data from gathering excel.....");
	sheets = wb.get_sheet_names()
	for sheet in sheets:		
		if sheet in sheetList :
			log("Extracting Data for " + sheet + " .....");			
			ws = wb.get_sheet_by_name(sheet)
			count = 1;
			while count < 6 :
				key = "RANGE" + str(count);				
				if(key not in bpt_config[sheet]):
					break		
				fileName = 	sheet.replace(" ","_")+"_" + key ;	
				fcsv = open(fileName + ".dat" ,"w");
				startRow = bpt_config[sheet][key][0];
				startCol = bpt_config[sheet][key][1];
				endRow = bpt_config[sheet][key][2];
				endCol = bpt_config[sheet][key][3];
				pullData(fcsv, ws, startRow, startCol, endRow, endCol);
				fcsv.close();
				bptDatList.append((fileName,"N"))
				count = count + 1
			
	log("Data extraction complete.....");

def pullData(fcsv, ws, startRow, startCol, endRow, endCol):	
	rowCount = colCount = 0;	
	for row in ws.rows:				
		rowCount = rowCount + 1;
		if rowCount > endRow :
			break;		
		if rowCount < startRow :
			continue;		
		colCount = 0
		eofL = False			
		dataStream = ""
		for cell in row:						
			colCount = colCount + 1
			if colCount > endCol:
				break;			
			if colCount < startCol:
				continue;			
			if ((cell.value) and cell.font.strikethrough):				
				break;
			value = ""
			if (str(cell.value) != 'None'):
				value = str(cell.value)			
			if ((cell.value or value == '0') and not cell.font.strikethrough):								
				dataStream = dataStream + str(cell.value) + ",";
				eofL = True;			
		if(eofL):
			fcsv.write(dataStream + "\n")			
		
		
	
def load():
	log("Laoding BPT data in table.....");
	for fileDet in bptDatList:
		cmd = "'C:\ora1107w\BIN\sqlldr.exe cntdb83/cntdb83@cntepcz control=" + fileDet[0] + ".ctl log=" + fileDet[0] + ".log bad=" + fileDet[0] + ".bad',shell=True"
		#subprocess.call(cmd)
		log (cmd)
	#subprocess.call('C:\ora1107w\BIN\sqlldr.exe cntdb83/cntdb83@cntepcz control=Bill_Cycles.ctl log=Bill_Cycles.log bad=Bill_Cycles.bad',shell=True)
	#subprocess.call('C:\ora1107w\BIN\sqlldr.exe cntdb83/cntdb83@cntepcz control=Messages_Entities.ctl log=Messages_Entities.log bad=Messages_Entities.bad',shell=True)
	#subprocess.call('C:\ora1107w\BIN\sqlldr.exe cntdb83/cntdb83@cntepcz control=Quality_Assurance.ctl log=Quality_Assurance.log bad=Quality_Assurance.bad',shell=True)

def validate():
	log("Validating BPT data.....");
	
def term():
	wb.close();
	close(logFile);

init();
#import pdb; pdb.set_trace()
extract(bl_wb,bl_sheets);
extract(cm_wb,cm_sheets);
extract(ar_wb,ar_sheets);
extract(cl_wb,cl_sheets);
load();
validate();

	
